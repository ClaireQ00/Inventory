#!/usr/bin/env python3
"""简要报价单导出 (2026-08-11, 模板 data/简要报价模板.xltx)

原则(老板定): 数值全部从系统取(录入时 db_writer 已算好落库), 模板只负责摆格式。
- 分组键 = (产品类别, 物料类型, 公斤价); 同类别两种公斤价 → 拆两组
- 卷价以库中 unit_price 为准(允许手填偏离 系数×单重, 不反算不改写)
- 行数不固定: 按实际明细动态行, 合计/底部自动下移, 格式复制模板行
用法: python3 tools/export_brief_quote.py QT20260812001 [--out output/quotes]
"""
import argparse
import os
from copy import copy
from pathlib import Path

import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Color

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "data" / "简要报价模板.xltx"
BLANK_SHEET = "「0728」 （空白模板） "
COLS = "ABCDEFGHIJKLMNO"  # A..O 15 列


def get_conn():
    pw = [l.split("=", 1)[1].strip() for l in open(ROOT / ".env") if l.startswith("MYSQL_PASSWORD=")][0]
    return pymysql.connect(host="127.0.0.1", port=3306, user="inventory", password=pw,
                           database="inventory_db", cursorclass=pymysql.cursors.DictCursor)


def fetch_quote(quote_no: str):
    conn = get_conn()
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM quotations WHERE quote_no=%s", (quote_no,))
        header = cur.fetchone()
        if not header:
            raise SystemExit(f"报价单不存在: {quote_no}")
        cur.execute(
            """SELECT i.*, p.product_category, p.material_type, p.inner_diameter, p.inner_diameter_inch,
                      p.outer_diameter, p.thickness, p.length,
                      p.appearance_inner, p.appearance_outer, p.appearance_height, p.volume AS p_volume
               FROM quotation_items i LEFT JOIN products p ON p.material_id = i.material_id
               WHERE i.quote_no=%s ORDER BY i.id""", (quote_no,))
        items = cur.fetchall()
    conn.close()
    return header, items


def group_items(items):
    """按 (产品类别, 物料类型, 公斤价) 分组, 保持出现顺序; 同类别不同公斤价拆两组"""
    groups, order = {}, []
    for it in items:
        key = (it["product_category"] or "未分类", it["material_type"] or "-",
               float(it["price_coefficient"]))
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(it)
    return [(k, groups[k]) for k in order]


def snap_style(ws, row):
    """抓取模板某行 A..O 的样式快照 (字体/边框/填充/对齐/数字格式)"""
    return {c: {"font": copy(ws[f"{c}{row}"].font), "border": copy(ws[f"{c}{row}"].border),
                "fill": copy(ws[f"{c}{row}"].fill), "alignment": copy(ws[f"{c}{row}"].alignment),
                "number_format": ws[f"{c}{row}"].number_format} for c in COLS}


def put(ws, row, styles, values, height=None, shrink_cols="", red_cols=""):
    for c in COLS:
        cell = ws[f"{c}{row}"]
        st = styles[c]
        font = copy(st["font"])
        if c in shrink_cols and font.size:
            font.size = max(font.size - 2, 6)  # 外观尺寸 4 列小两号 (老板 2026-08-11)
        if c in red_cols:
            font.color = Color(rgb="FFFF0000")  # 合计金额标红 (老板 2026-08-11)
        cell.font, cell.border, cell.fill, cell.alignment = font, st["border"], st["fill"], st["alignment"]
        cell.number_format = st["number_format"]
        if c in values:
            cell.value = values[c]
    if height:
        ws.row_dimensions[row].height = height


def export(quote_no: str, out_dir: Path) -> Path:
    header, items = fetch_quote(quote_no)
    groups = group_items(items)

    wb = load_workbook(TEMPLATE)
    ws = wb[BLANK_SHEET]
    ws.title = quote_no
    # 只留空白模板这一张
    for name in list(wb.sheetnames):
        if name != quote_no:
            del wb[name]

    # 样式快照: 组头(2) / 列头(3) / 明细(4) / 合计(12) / 合计数+付款(13) / 条款(14)
    st_group, st_col, st_item = snap_style(ws, 2), snap_style(ws, 3), snap_style(ws, 4)
    st_total, st_sum, st_terms = snap_style(ws, 12), snap_style(ws, 13), snap_style(ws, 14)

    # 清空 2 行起的全部内容与合并
    for rng in list(ws.merged_cells.ranges):
        if str(rng) != "A1:K1":
            ws.unmerge_cells(str(rng))
    ws.delete_rows(2, ws.max_row)

    r = 2
    seq = 0
    tot_rolls, tot_cbm, tot_kg = 0, 0.0, 0.0
    for (cat, mtype, coeff), rows in groups:
        put(ws, r, st_group, {"A": f"{cat}——{mtype}", "H": coeff,
                              "I": "Subtotal 小计", "L": "预估外观尺寸 Apperance Estimate"}, height=27)
        for rng in (f"A{r}:G{r}", f"I{r}:K{r}", f"L{r}:O{r}"):
            ws.merge_cells(rng)
        r += 1
        put(ws, r, st_col, {"A": "No.", "B": "标称英寸\nInch", "C": "内径\nID(mm)", "D": "外径\nOD(mm)",
                            "E": "厚\n(mm)", "F": "长度\nM/roll", "G": "重量\nKG/ROLL",
                            "H": f"卷价\nPRICE/ROLL({header['currency']})", "I": "卷数\nROLLS",
                            "J": "体积\nCBM", "K": "总重\nKG", "L": "内圈\nInner", "M": "外圈\nOutter",
                            "N": "高度\nHeight", "O": "单卷体积\nCBM/ROLL"}, height=40)
        r += 1
        for it in rows:
            seq += 1
            qty = int(it["quantity"])
            wt = float(it["weight_per_unit"])
            vol = float(it["volume"] or it["p_volume"] or 0)
            tot_rolls += qty
            tot_cbm += vol * qty
            tot_kg += wt * qty
            put(ws, r, st_item, {
                "A": seq, "B": it["inner_diameter_inch"], "C": it["inner_diameter"],
                "D": it["outer_diameter"], "E": it["thickness"], "F": it["length"],
                "G": wt, "H": float(it["unit_price"]), "I": qty,
                "J": round(vol * qty, 4), "K": round(wt * qty, 3),
                "L": it["appearance_inner"], "M": it["appearance_outer"],
                "N": it["appearance_height"], "O": vol}, height=20, shrink_cols="LMNO")
            r += 1

    # 合计行: 卷数/体积/重量/金额
    put(ws, r, st_total, {"A": "合计 TOTAL"}, height=20)
    ws.merge_cells(f"A{r}:K{r}")
    r += 1
    put(ws, r, st_sum, {"A": "付款条件：", "B": header["payment_term"] or "",
                        "H": float(header["total_amount"]),  # 金额合计(系统落库值, 标红)
                        "I": tot_rolls, "J": round(tot_cbm, 2), "K": round(tot_kg, 1),
                        "L": header["packing"] or ""}, height=49, red_cols="H")
    ws.merge_cells(f"B{r}:G{r}")  # 付款条件长文本合并显示 (老板 2026-08-11)
    ws.merge_cells(f"L{r}:O{r}")
    for cc in (f"B{r}", f"L{r}"):  # 长文本开自动换行
        al = copy(ws[cc].alignment)
        al.wrap_text = True
        al.vertical = "center"
        ws[cc].alignment = al
    r += 1
    delivery = f"收到定金后{header['delivery_days']}天" if header.get("delivery_days") else ""
    trade = f"{header['trade_terms']} {header['port_loading'] or ''}".strip()
    put(ws, r, st_terms, {"A": "报价方式：", "B": trade, "D": "交货期：", "E": delivery,
                          "H": "有效期至：", "I": header["valid_until"], "K": header["customer_code"]},
        height=40)
    ws.merge_cells(f"B{r}:C{r}")  # 报价方式 (如 FOB Qingdao)
    ws.merge_cells(f"E{r}:G{r}")  # 交货期
    ws.merge_cells(f"I{r}:J{r}")  # 有效期日期合并两格才显示全 (老板 2026-08-11)
    for cc in (f"B{r}", f"E{r}"):
        al = copy(ws[cc].alignment)
        al.wrap_text = True
        al.vertical = "center"
        ws[cc].alignment = al
    ws[f"I{r}"].number_format = "yyyy/m/d"

    # 标题
    ws["A1"] = f"YONGLI PVC HOSE 永利软管简要报价——{quote_no}——{header['currency']}"

    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{quote_no}_简要报价.xlsx"
    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("quote_no")
    ap.add_argument("--out", default=str(ROOT / "output" / "quotes"))
    a = ap.parse_args()
    print(export(a.quote_no, Path(a.out)))


if __name__ == "__main__":
    main()
