#!/usr/bin/env python3
"""gen_distribution_table.py — 从数据库生成中英文 GOODS DISTRIBUTION TABLE 三件套
  1. GOODS DISTRIBUTION TABLE (货物分布总表) — 已装柜/临沂库存/本厂库存 三态分布
  2. TRANSFER ORDER 调拨单 (本厂→临沂)
  3. LOADING LIST 装柜出库单 (临沂→海运)
数量全部来自 stock_logs/inventory/合同, 不手填。

当前为 BL-2608 (Q025 印尼大雄) 实例: SPECS/合同号/日期按该批次写死。
第二个订单来时把顶部常量抽成命令行参数即可通用化。
运行: MYSQL_PASSWORD=xxx python3 tools/gen_distribution_table.py
产物: output/BL-2608/BL-2608_出入库表单_系统生成.xlsx (output/ 已 gitignore)
"""
import os
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = dict(host="127.0.0.1", port=3306, user=os.environ.get("DB_USER", "inventory"),
          password=os.environ["MYSQL_PASSWORD"], database=os.environ.get("DB_NAME", "inventory_db"),
          charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor)
conn = pymysql.connect(**DB)

SPECS = ["M-Q025-001", "M-Q025-002", "M-Q025-014", "M-Q025-004",
         "M-Q025-006", "M-Q025-013", "M-Q025-011"]  # 按英寸从小到大

# 历史货物无系统合同价, 用原报表实测价 (备注注明); 013 体积缺失用原报表 0.13
PRICE_FALLBACK = {"M-Q025-014": 10.70, "M-Q025-011": 55.22}
VOLUME_FALLBACK = {"M-Q025-013": 0.13}

with conn.cursor() as cur:
    cur.execute("""SELECT material_id, spec, inner_diameter_inch, weight, volume
                   FROM products WHERE material_id IN %s""", (tuple(SPECS),))
    prods = {r["material_id"]: r for r in cur.fetchall()}
    # 合同价 (730 + 802)
    cur.execute("""SELECT material_id, unit_price FROM sales_contract_items
                   WHERE contract_no IN ('SC20260730001','SC20260802001')""")
    prices = {r["material_id"]: float(r["unit_price"]) for r in cur.fetchall()}
    # ① 装柜 = 0808 临沂全部出库 (销售+调整)
    cur.execute("""SELECT soi.material_id, SUM(soi.quantity) AS qty FROM stock_out_items soi
                   JOIN stock_out so ON so.out_no=soi.out_no
                   WHERE so.warehouse_code='WH-03' AND so.out_date='2026-08-08' AND so.status='confirmed'
                   GROUP BY soi.material_id""")
    shipped = {r["material_id"]: int(r["qty"]) for r in cur.fetchall()}
    # ②③ 当前库存
    cur.execute("""SELECT warehouse_code, material_id, quantity FROM inventory
                   WHERE material_id IN %s""", (tuple(SPECS),))
    stock = {(r["warehouse_code"], r["material_id"]): int(r["quantity"]) for r in cur.fetchall()}
    # 调拨明细 (0807)
    cur.execute("""SELECT soi.material_id, soi.quantity FROM stock_out_items soi
                   JOIN stock_out so ON so.out_no=soi.out_no
                   WHERE so.transfer_ref='TR20260807001' AND so.out_type='transfer'""")
    transfer = list(cur.fetchall())
    # 装柜分明细来源: 销售(合同锚定) vs 调整(历史货物)
    cur.execute("""SELECT so.out_type, soi.material_id, soi.quantity FROM stock_out_items soi
                   JOIN stock_out so ON so.out_no=soi.out_no
                   WHERE so.warehouse_code='WH-03' AND so.out_date='2026-08-08' AND so.status='confirmed'
                   ORDER BY so.out_type DESC, soi.material_id""")
    loading = list(cur.fetchall())

def vol(mid):
    v = prods[mid]["volume"]
    return float(v) if v is not None else VOLUME_FALLBACK.get(mid, 0.0)

def price(mid):
    return prices.get(mid) or PRICE_FALLBACK.get(mid)

# ── 样式 ──
thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
TOT_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_row(ws, r, cols, fill=None, bold=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = CENTER
        if fill:
            cell.fill = fill
        if bold:
            cell.font = BOLD

# ═══ 表1: GOODS DISTRIBUTION TABLE ═══
wb = Workbook()
ws = wb.active
ws.title = "GOODS DISTRIBUTION TABLE"
ws.merge_cells("A1:L1")
ws["A1"] = "GOODS DISTRIBUTION TABLE — Contract SC20260730001 + SC20260802001  货物分布总表"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A2:L2")
ws["A2"] = "Product: WATER HOSE (orange)  ·  Customer: Q025 印尼大雄  ·  B/L: BL-2608  ·  Date 日期: 2026-08-08"
headers = ["#\n序号", "Specification\n规格", "Unit Price\n单价\nUSD/roll", "Weight\n重量\nkg/roll",
           "CBM\n体积\nm³/roll",
           "① Shipped 已装柜\nRolls 卷", "① CBM m³",
           "② In Stock Linyi 临沂库存\nRolls 卷", "② CBM m³",
           "③ In Stock Factory 本厂库存\nRolls 卷", "③ CBM m³",
           "Total\nRolls 总卷数"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_row(ws, 4, 12, HEAD_FILL, True)

r = 5
tot = [0, 0.0, 0, 0.0, 0, 0.0, 0]
for i, mid in enumerate(SPECS, 1):
    p = prods[mid]
    sh, ly, fc = shipped.get(mid, 0), stock.get(("WH-03", mid), 0), stock.get(("WH-01", mid), 0)
    v = vol(mid)
    row = [i, p["spec"], price(mid), float(p["weight"]), round(v, 4),
           sh, round(sh * v, 3), ly, round(ly * v, 3), fc, round(fc * v, 3), sh + ly + fc]
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
    style_row(ws, r, 12)
    tot[0] += sh; tot[1] += sh * v; tot[2] += ly; tot[3] += ly * v
    tot[4] += fc; tot[5] += fc * v; tot[6] += sh + ly + fc
    r += 1
ws.cell(row=r, column=1, value="TOTAL")
ws.cell(row=r, column=2, value="合计")
for c, val in [(6, tot[0]), (7, round(tot[1], 3)), (8, tot[2]), (9, round(tot[3], 3)),
               (10, tot[4]), (11, round(tot[5], 3)), (12, tot[6])]:
    ws.cell(row=r, column=c, value=val)
style_row(ws, r, 12, TOT_FILL, True)
notes = [
    "NOTES & REMARKS 备注:",
    "1. Quantities are system-generated from stock records (stock_in/stock_out/stock_logs), not manual entry. 数量由系统出入库流水自动生成。",
    "2. ① Shipped = 2026-08-08 Linyi loading (511 contract-anchored + 280 legacy opening stock). ①=0808临沂装柜(合同锚定511卷+历史期初货物280卷)。",
    "3. ③ Factory stock 11 rolls of 1/4\" = pre-system manual entry on 2026-08-02. 本厂1/4寸11卷为08-02既有录入。",
    "4. Unit price: contract price where available; M-Q025-014/011 use measured price from source sheet (no contract in system). 单价优先取合同价, 3/4寸与2寸无系统合同, 取原报表实测价。",
    "5. CBM/roll from products.volume (OD²×H×0.93 rule, appearance dims updated to BL-2608 batch measurements 2026-08-10). 单卷体积取物料主数据(0.93装箱系数), 外观尺寸已按本批实测更新。",
]
for i, t in enumerate(notes):
    ws.cell(row=r + 2 + i, column=1, value=t).alignment = Alignment(wrap_text=False)
for i, w in enumerate([5, 22, 11, 10, 10, 11, 9, 12, 9, 12, 9, 9], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 52

# ═══ 表2: TRANSFER ORDER 调拨单 ═══
ws2 = wb.create_sheet("TRANSFER ORDER 调拨单")
ws2.merge_cells("A1:H1")
ws2["A1"] = "STOCK TRANSFER ORDER  调拨单  TR20260807001"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A2:H2")
ws2["A2"] = "From 调出: WH-01 南线仓库(本厂)  →  To 调入: WH-03 临沂赛君仓库  ·  Date 日期: 2026-08-07  ·  Docs 单号: OUT20260807001 / IN20260807003"
h2 = ["#\n序号", "Material 物料号", "Specification 规格", "Qty 数量\nRolls 卷", "Weight 单重\nkg/roll",
      "Total Weight 总重\nkg", "CBM 单卷体积\nm³", "Total CBM 总体积\nm³"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=4, column=i, value=h)
style_row(ws2, 4, 8, HEAD_FILL, True)
r = 5
tq = tk = tv = 0
for i, it in enumerate(transfer, 1):
    mid = it["material_id"]
    p = prods[mid]
    q = int(it["quantity"])
    row = [i, mid, p["spec"], q, float(p["weight"]), round(q * float(p["weight"]), 1),
           round(vol(mid), 4), round(q * vol(mid), 3)]
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 8)
    tq += q; tk += q * float(p["weight"]); tv += q * vol(mid)
    r += 1
ws2.cell(row=r, column=1, value="TOTAL")
ws2.cell(row=r, column=2, value="合计")
for c, val in [(4, tq), (6, round(tk, 1)), (8, round(tv, 3))]:
    ws2.cell(row=r, column=c, value=val)
style_row(ws2, r, 8, TOT_FILL, True)
ws2.cell(row=r + 2, column=1, value="备注: 本厂出库=临沂入库, 同一调拨号 TR20260807001 配对, 校验第14步自动核对两边数量一致。")
for i, w in enumerate([5, 14, 22, 10, 10, 12, 11, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[4].height = 40

# ═══ 表3: LOADING LIST 装柜出库单 ═══
ws3 = wb.create_sheet("LOADING LIST 装柜单")
ws3.merge_cells("A1:H1")
ws3["A1"] = "LOADING LIST  装柜出库单  BL-2608"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A2:H2")
ws3["A2"] = "Warehouse 出库仓: WH-03 临沂中转仓  →  Sea freight 海运  ·  Date 日期: 2026-08-08  ·  Docs 单号: OUT20260808001(销售) / OUT20260808002(调整) / DN20260808001(发货单)"
for i, h in enumerate(["#\n序号", "Material 物料号", "Specification 规格", "Qty 数量\nRolls 卷",
                       "Source 来源", "Total Weight 总重\nkg", "Total CBM 总体积\nm³", "Remark 备注"], 1):
    ws3.cell(row=4, column=i, value=h)
style_row(ws3, 4, 8, HEAD_FILL, True)
r = 5
tq = tk = tv = 0
src_map = {"sale": "合同锚定 Contract", "adjust": "历史期初货物 Legacy"}
for i, it in enumerate(loading, 1):
    mid = it["material_id"]
    p = prods[mid]
    q = int(it["quantity"])
    w = float(p["weight"])
    row = [i, mid, p["spec"], q, src_map.get(it["out_type"], it["out_type"]),
           round(q * w, 1), round(q * vol(mid), 3), ""]
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, 8)
    tq += q; tk += q * w; tv += q * vol(mid)
    r += 1
ws3.cell(row=r, column=1, value="TOTAL")
ws3.cell(row=r, column=2, value="合计")
for c, val in [(4, tq), (6, round(tk, 1)), (7, round(tv, 3))]:
    ws3.cell(row=r, column=c, value=val)
style_row(ws3, r, 8, TOT_FILL, True)
ws3.cell(row=r + 2, column=1, value="备注: 合同锚定511卷=SC20260730001全额500卷(纸面)+SC20260802001首批11卷(M-Q025-013); 历史期初货物280卷=0726临沂结余+1寸超产4卷。")
ws3.cell(row=r + 3, column=1, value="注: SC20260730001第5行纸面码M-Q025-012, 实物为M-Q025-013(ID42), 已在销售出库按实物出。")
for i, w in enumerate([5, 14, 22, 10, 18, 12, 12, 10], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[4].height = 40

os.makedirs("output/BL-2608", exist_ok=True)
path = "output/BL-2608/BL-2608_出入库表单_系统生成.xlsx"
wb.save(path)
print("saved:", path)
conn.close()
