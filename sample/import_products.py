#!/usr/bin/env python3
"""Generate SQL INSERT statements for products from an Excel template."""

from __future__ import annotations

import argparse
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PRODUCT_FIELDS = [
    'material_id',
    'customer_code',
    'brand',
    'product_category',
    'material_type',
    'spec',
    'inner_diameter',
    'inner_diameter_inch',
    'outer_diameter',
    'id_x_od',
    'thickness',
    'length',
    'virtual_weight',
    'virtual_length',
    'wire_spacing',
    'weight_per_meter',
    'weight',
    'appearance_inner',
    'appearance_outer',
    'appearance_height',
    'volume',
    'volume_subtotal',
    'package',
    'label_paper',
    'material_used',
    'wire_pattern',
    'coil_type',
    'pressure',
    'spray_code',
    'meter_mark',
    'remark',
    'is_active',
]

DEFAULT_VALUES = {
    'is_active': 1,
}

NUMERIC_FIELDS = {
    'inner_diameter',
    'outer_diameter',
    'thickness',
    'length',
    'virtual_weight',
    'virtual_length',
    'weight_per_meter',
    'weight',
    'appearance_inner',
    'appearance_outer',
    'appearance_height',
    'volume',
    'volume_subtotal',
    'pressure',
}

REQUIRED_FIELDS = {'material_id', 'spec'}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='从 Excel 模板生成 products 表的 INSERT SQL 语句。'
    )
    parser.add_argument('xlsx_path', type=Path, help='Excel 模板文件路径')
    parser.add_argument(
        '--sheet', default='products', help='包含物料数据的工作表名（默认: products）'
    )
    parser.add_argument(
        '--output', default='sample/import_products.sql', help='输出 SQL 文件路径'
    )
    return parser.parse_args()


def format_decimal(value: Any) -> str:
    decimal_value = Decimal(str(value))
    normalized = decimal_value.normalize()
    text = format(normalized, 'f')
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return text or '0'


def normalize_value(field: str, raw: Any) -> Any:
    if raw is None:
        return None
    if isinstance(raw, str):
        raw = raw.strip()
        if raw == '':
            return None
        return raw
    if isinstance(raw, bool):
        return int(raw)
    if field in NUMERIC_FIELDS:
        try:
            return Decimal(str(raw))
        except Exception:
            return raw
    return raw


def sql_literal(value: Any) -> str:
    if value is None:
        return 'NULL'
    if isinstance(value, Decimal):
        return format_decimal(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format_decimal(Decimal(str(value)))
    text = str(value)
    text = text.replace("'", "''")
    return f"'{text}'"


def compute_derived_fields(row_data: dict[str, Any]) -> None:
    if row_data.get('outer_diameter') is None:
        inner = row_data.get('inner_diameter')
        thickness = row_data.get('thickness')
        if inner is not None and thickness is not None:
            try:
                row_data['outer_diameter'] = inner + thickness * Decimal('2')
            except Exception:
                pass
    if row_data.get('id_x_od') is None:
        inner = row_data.get('inner_diameter')
        outer = row_data.get('outer_diameter')
        if inner is not None and outer is not None:
            row_data['id_x_od'] = f"{format_decimal(inner)}x{format_decimal(outer)}"
    # 单件体积 = 外观外径(mm)² × 外观高度(mm) × 0.93 / 1e6  (圆盘装箱经验系数)
    if row_data.get('volume') is None:
        ao = row_data.get('appearance_outer')
        ah = row_data.get('appearance_height')
        if ao is not None and ah is not None:
            try:
                vol = (Decimal(str(ao)) ** 2) * Decimal(str(ah)) * Decimal('0.93') / Decimal('1000000')
                row_data['volume'] = vol.quantize(Decimal('0.0001'))
            except Exception:
                pass


def read_products_from_excel(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError('模板中找不到表头行。')
    header = [str(col).strip() if col is not None else '' for col in first_row]
    while header and header[-1] == '':
        header.pop()
    if header != PRODUCT_FIELDS:
        raise ValueError(
            '模板表头与预期字段不匹配。\n'
            f'预期字段: {PRODUCT_FIELDS}\n'
            f'实际字段: {header}'
        )

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        values = list(row)
        if len(values) < len(PRODUCT_FIELDS):
            values.extend([None] * (len(PRODUCT_FIELDS) - len(values)))
        row_data = {field: normalize_value(field, value) for field, value in zip(PRODUCT_FIELDS, values)}
        if not row_data.get('material_id'):
            continue
        compute_derived_fields(row_data)
        for field, default in DEFAULT_VALUES.items():
            if row_data.get(field) is None:
                row_data[field] = default
        rows.append(row_data)
    return rows


def build_insert_sql(rows: list[dict[str, Any]]) -> str:
    columns = ', '.join(PRODUCT_FIELDS)
    items = []
    for index, row in enumerate(rows, start=1):
        missing = [field for field in REQUIRED_FIELDS if not row.get(field)]
        if missing:
            raise ValueError(f'第 {index + 1} 行缺少必填字段: {missing}')
        values = ', '.join(sql_literal(row.get(field)) for field in PRODUCT_FIELDS)
        items.append(f'INSERT INTO products ({columns}) VALUES ({values});')
    return '\n'.join(items)


def main() -> None:
    args = parse_args()
    products = read_products_from_excel(args.xlsx_path, args.sheet)
    if not products:
        raise SystemExit('未找到可导入的物料记录，请检查模板是否已填写数据。')

    sql_text = build_insert_sql(products)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql_text, encoding='utf-8')
    print(f'已生成 {len(products)} 条 INSERT 语句到: {output_path}')
    print('请使用 MySQL 执行生成的 SQL 文件。')


if __name__ == '__main__':
    main()
